import json
import os
from collections import defaultdict

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

file_name = "mmcl_cnn_4layer_b_rvcl_cnn_4layer_b_adv_regular_cl_cnn_4layer_b_supervised_cnn_4layer_b"
with open(f"../rs_results/{file_name}.json", "r") as f:
    data = json.load(f)

sigma_values = [0.25, 0.5, 1]
certified_radius_choices = [0, 0.5, 1, 1.5, 2, 2.5, 3]
model_names = ["mmcl", "rvcl", "regular_cl", "supervised"]
per_model = defaultdict(list)
per_sigma_radius = defaultdict(list)

for model in model_names:
    for curr_sigma in sigma_values:
        # Get all examples for the current sigma value (denominator remains constant)
        all_values = [x for x in data[model] if x["sigma"] == curr_sigma]
        total = len(all_values)
        for curr_radius in certified_radius_choices:
            certified_count = sum(
                1 for x in all_values
                if x["radius"] >= curr_radius and x["true_label"] == x["rs_label"]
            )
            unchanged_count = sum(
                1 for x in all_values
                if x["radius"] >= curr_radius and x["predicted_label"] == x["rs_label"]
            )
            certified_accuracy = certified_count / total if total > 0 else 0
            unchanged_percentage = unchanged_count / total if total > 0 else 0

            per_model[model].append({
                "sigma": curr_sigma,
                "radius": curr_radius,
                "certified_accuracy": certified_accuracy,
                "unchanged_percentage": unchanged_percentage
            })
            per_sigma_radius[f"sigma:{curr_sigma}|radius:{curr_radius}"].append({
                "model": model,
                "certified_accuracy": certified_accuracy,
                "unchanged_percentage": unchanged_percentage
            })

per_sigma_radius_updated = {}
for key, values in per_sigma_radius.items():
    max_certified_accuracy = max(x["certified_accuracy"] for x in values)
    best_certified_accuracy_model = (
        [x["model"] for x in values if x["certified_accuracy"] == max_certified_accuracy]
        if max_certified_accuracy > 0
        else []
    )
    max_unchanged_percentage = max(x["unchanged_percentage"] for x in values)
    best_unchanged_percentage_model = (
        [x["model"] for x in values if x["unchanged_percentage"] == max_unchanged_percentage]
        if max_unchanged_percentage > 0
        else []
    )
    per_sigma_radius_updated[key] = {
        "per_model_values": values,
        "best_certified_accuracy_models": best_certified_accuracy_model,
        "best_unchanged_percentage_models": best_unchanged_percentage_model
    }

print(json.dumps(per_sigma_radius_updated, indent=4))

# Prepare table rows for Excel
rows = []
for key, value in per_sigma_radius_updated.items():
    sigma_str, radius_str = key.split("|")
    sigma = float(sigma_str.split(":")[1])
    radius = float(radius_str.split(":")[1])
    per_model_values = value["per_model_values"]
    # Create a mapping from model name to its metrics
    model_metrics = {x["model"]: x for x in per_model_values}
    row = {
        "sigma": sigma,
        "radius": radius,
    }
    for model in model_names:
        metrics = model_metrics[model]
        row[f"{model}_cia"] = round(metrics["certified_accuracy"], 2)
        row[f"{model}_up"] = round(metrics["unchanged_percentage"], 2)
    row["best_certified_model"] = ", ".join(value["best_certified_accuracy_models"])
    row["best_unchanged_model"] = ", ".join(value["best_unchanged_percentage_models"])
    rows.append(row)

df = pd.DataFrame(rows)
df = df.sort_values(by=["sigma", "radius"]).reset_index(drop=True)

excel_filename = f"../rs_results/{file_name}.xlsx"
df.to_excel(excel_filename, index=False)

# Load workbook and apply per-row bold formatting
wb = load_workbook(excel_filename)
ws = wb.active

# Get header names to identify columns for certified accuracy (_cia) and unchanged percentage (_up)
header_names = [cell.value for cell in ws[1]]
cia_indices = [i+1 for i, name in enumerate(header_names) if isinstance(name, str) and name.endswith("_cia")]
up_indices = [i+1 for i, name in enumerate(header_names) if isinstance(name, str) and name.endswith("_up")]

# Iterate over each data row
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    current_row = row[0].row  # get current row number
    # Process certified accuracy columns
    cia_values = [(ws.cell(row=current_row, column=col).value, col) for col in cia_indices]
    # Filter numeric values and determine maximum for certified accuracy
    cia_numeric = [value for value, _ in cia_values if isinstance(value, (int, float))]
    if cia_numeric:
        max_cia = max(cia_numeric)
        for value, col in cia_values:
            if isinstance(value, (int, float)) and value == max_cia:
                ws.cell(row=current_row, column=col).font = Font(bold=True)
    # Process unchanged percentage columns
    up_values = [(ws.cell(row=current_row, column=col).value, col) for col in up_indices]
    up_numeric = [value for value, _ in up_values if isinstance(value, (int, float))]
    if up_numeric:
        max_up = max(up_numeric)
        for value, col in up_values:
            if isinstance(value, (int, float)) and value == max_up:
                ws.cell(row=current_row, column=col).font = Font(bold=True)

wb.save(excel_filename)
print(f"Excel file with per-row formatting saved to: {os.path.abspath(excel_filename)}")


def plot_one_per_sigma(data):
    model_names = ["mmcl", "rvcl", "regular_cl", "supervised"]
    sigma_set = set()
    for model in model_names:
        for rec in data[model]:
            sigma_set.add(rec["sigma"])
    sigma_values = sorted(list(sigma_set))
    for sigma in sigma_values:
        plt.figure(figsize=(8,6))
        max_threshold = 0
        model_records = {}
        for model in model_names:
            records = [r for r in data[model] if r["sigma"] == sigma]
            model_records[model] = records
            if records:
                max_val = max(r["radius"] for r in records)
                if max_val > max_threshold:
                    max_threshold = max_val
        if max_threshold == 0:
            max_threshold = 3.5
        x_vals = np.linspace(0, max_threshold, 200)
        for model in model_names:
            records = model_records[model]
            total = len(records)
            y_vals = []
            for r in x_vals:
                count = sum(1 for rec in records if rec["radius"] >= r and rec["true_label"] == rec["rs_label"])
                y_vals.append(count / total if total > 0 else 0)
            plt.plot(x_vals, y_vals, label=model)
        plt.xlabel("Radius Threshold")
        plt.ylabel("Certified Accuracy")
        plt.title(f"Certified Accuracy vs Radius (sigma = {sigma})")
        plt.legend()
        plt.grid(True)
        output_filename = f"per_sigma_comparison_sigma_{sigma}.png"
        plt.tight_layout()
        plt.savefig(os.path.join("../rs_results", output_filename))
        plt.close()
        print(f"Plot saved as: {output_filename}")

plot_one_per_sigma(data)